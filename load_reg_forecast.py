#!/usr/bin/env python3
"""
load_reg_forecast.py — Read Q2 Apollo unit forecast from spreadsheet and
merge apolloRegForecast into data.json.

Source: ~/Downloads/Q2 2026 Weekly Reg Forecast.xlsx  (Apollo sheet)
Rows used:
  row 4  — APOLLO DESKTOP  / 2026 Forecast  (cols B–N)
  row 9  — APOLLO RACKMOUNT / 2026 Forecast (cols B–N)
  row 14 — COMBINED TOTAL  / 2026 Forecast  (cols B–N)
"""
import json
import sys
from pathlib import Path

XLSX = Path.home() / "Downloads" / "Q2 2026 Weekly Reg Forecast.xlsx"
DATA_JSON = Path(__file__).parent / "data.json"

WEEK_ISO = {
    "Apr 06": "2026-04-06", "Apr 13": "2026-04-13",
    "Apr 20": "2026-04-20", "Apr 27": "2026-04-27",
    "May 04": "2026-05-04", "May 11": "2026-05-11",
    "May 18": "2026-05-18", "May 25": "2026-05-25",
    "Jun 01": "2026-06-01", "Jun 08": "2026-06-08",
    "Jun 15": "2026-06-15", "Jun 22": "2026-06-22",
    "Jun 29": "2026-06-29",
}


def main():
    if not XLSX.exists():
        print(f"  WARNING: {XLSX} not found — skipping apolloRegForecast")
        return

    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed — skipping apolloRegForecast")
        print("  Install with: pip3 install openpyxl")
        return

    wb = openpyxl.load_workbook(str(XLSX))
    if "Apollo" not in wb.sheetnames:
        print("  WARNING: 'Apollo' sheet not found in spreadsheet — skipping")
        return

    ws = wb["Apollo"]
    rows = list(ws.iter_rows(values_only=True))

    # Row index 1 = header row with week labels ("Apr 06", ...)
    header = rows[1]
    week_labels = [str(h) for h in header[1:14]]
    weeks_iso = [WEEK_ISO.get(lbl) for lbl in week_labels]
    if None in weeks_iso:
        print(f"  WARNING: unexpected week label(s) in header: {week_labels}")
        weeks_iso = [w for w in weeks_iso if w]

    # Row index 3 = Desktop 2026 Forecast
    # Row index 8 = Rackmount 2026 Forecast
    # Row index 13 = Combined 2026 Forecast
    desktop   = [int(v) if v is not None else 0 for v in rows[3][1:14]]
    rackmount = [int(v) if v is not None else 0 for v in rows[8][1:14]]
    combined  = [int(v) if v is not None else 0 for v in rows[13][1:14]]

    data = json.loads(DATA_JSON.read_text())
    data["apolloRegForecast"] = {
        "weeks":     weeks_iso,
        "desktop":   desktop,
        "rackmount": rackmount,
        "combined":  combined,
    }
    DATA_JSON.write_text(json.dumps(data, separators=(",", ":")))
    print(f"  apolloRegForecast loaded: {len(weeks_iso)} weeks")
    print(f"  Desktop Q2 total:   {sum(desktop):,}")
    print(f"  Rackmount Q2 total: {sum(rackmount):,}")
    print(f"  Combined Q2 total:  {sum(combined):,}")


if __name__ == "__main__":
    main()
